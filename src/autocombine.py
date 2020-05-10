#coding=utf-8

# date      :    2020/5/10 12:04
# Author    :    qimi
# Filename  :    autocombine.py

import logging
import datetime
import os
import openpyxl

# 将该表的数据汇总到另一张表
excelFromFile = r'C:\Users\Administrator\Desktop\excel\book2.xlsx'
# 将其他表的数据汇总到这个表
excelToFile = r'C:\Users\Administrator\Desktop\excel\book1.xlsx'
# 如果该列有指定则仅更新此列数据
referColumn = ['G', '张三']
# 需要更新的数据列
combineColumn = ['H', 'I']
# 日志默认文件名
logFileDefaultName = 'option.log'

# 工作表类
class excel:
    # 文件名
    fileName = ''
    # 打开的工作簿
    workBook = ''
    # 当前活动的工作表
    currentSheet = ''
    # 是否被打开
    active = False

    def __init__(self, fileName):
        self.fileName = fileName

    def open(self, sheetName):
        try:
            self.workBook = openpyxl.load_workbook(self.fileName)
            self.currentSheet = self.workBook[sheetName]
        except Exception as err:
            # exType, exValue, exTrace = sys.exc_info()
            # logging.error(str(exType) + str(exValue))
            logging.error(err)
            return False
        self.active = True
        return True


    def close(self):
        self.workBook.save(self.fileName)
        self.active = False

    # 设置当前活动的工作表
    def SetCurrentSheet(self, sheetName):
        self.currentSheet = self.workBook[sheetName]

    # 获取最大行和列
    def GetMaxRowAndColumn(self):
        maxRow = self.currentSheet.max_row
        maxColumn = self.currentSheet.max_column
        return maxRow, maxColumn

    # 将另一个excel表的指定列合并到指定列
    def Combine(self, excelFrom, referColumn, combineColumn):
        # 参数检查
        #比较两张表的行列匹配
        maxRowColumnFrom = excelFrom.GetMaxRowAndColumn()
        maxRowColumnTo = self.GetMaxRowAndColumn()
        if maxRowColumnFrom != maxRowColumnTo:
            logging.error("from(%d , %d) is not equal to(%d , %d) row and column,"\
                          %(maxRowColumnFrom[0], maxRowColumnFrom[1], \
                            maxRowColumnTo[0], maxRowColumnTo[1]))
            return False

        # 比较要合并的列是否超过表的最大列

        # 开始合入
        for i in range(maxRowColumnTo[0]):
            # logging.debug(type(referColumn[0]))
            match = self.currentSheet.cell(row=i+1, column=referColumn[0]).value
            if match != referColumn[1]:
                continue
            for column in combineColumn:
                value = excelFrom.currentSheet.cell(row=i+1, column=column).value
                self.currentSheet.cell(row=i+1, column=column).value = value
            # 写入一行就保存文件
            self.workBook.save(self.fileName)
            logging.debug("combine row : %d" %(i+1))

        # 合入完成
        logging.debug('combine finish')

# 将字母列转换为数字列
def ConvertExcelStrToNumColumn(column):
    i = 0
    for o in column:
        column[i] = openpyxl.utils.column_index_from_string(o)
        i += 1
def GetExcelColumnNumFromStr(str):
    return openpyxl.utils.column_index_from_string(str)

def ConvertReferColumn(column):
    column[0] = openpyxl.utils.column_index_from_string(column[0])

# 日志配置
def logConfig(level, iscover = True):
    logFileName = logFileDefaultName
    if iscover:
        # 如果文件存在则删除
        if os.path.exists(logFileName):
            os.remove(logFileName)
    else:
        time = datetime.datetime.now()
        logFileName = "%02d%02d%02d_%02d%02d%02d.%03d"\
                      %(time.year, time.month, time.day,\
                        time.hour, time.minute, time.second,\
                        time.microsecond) + '.log'
    format = '%(asctime)s -%(levelname)s- %(message)s'
    logging.basicConfig(filename=logFileName, level=level, format=format)

# 获取参数
def GetParamFromKeyBoard(fromFile, toFile, refer, combine):
    str = ''
    while True:
        fromFile = input('请输入需要合入的文件名(*.xlsx):')
        if not os.path.exists(fromFile):
            print('文件不存在,请重新输入')
            continue
        break
    while True:
        toFile = input('请输入合入文件名:')
        if not os.path.exists(toFile):
            print('文件不存在,请重新输入')
            continue
        break
    while True:
        str = input('请输入参考列(G,张三):')
        list = str.split(',')
        if len(list) != 2:
            print('参考列输入有误请重新输入:')
            continue
        refer.clear()
        for o in list:
            refer.append(o)
        break
    str = input('请输入需要合入的列(H,I,J):')
    combine.clear()
    for o in str.split(','):
        combine.append(o)


if __name__ == '__main__':
    ret = ''
    # 获取参数
    GetParamFromKeyBoard(excelFromFile, excelToFile, referColumn, combineColumn)
    # print(combineColumn)
    # 日志记录配置
    logConfig(logging.DEBUG)
    logging.debug("start autocombine excel ".center(30, '-'))
    while True:
        # 1.获取参数
        # 2.校验参数
        ConvertExcelStrToNumColumn(combineColumn)
        ConvertReferColumn(referColumn)
        logging.debug("combineColumn:" + str(combineColumn))
        excelFrom = excel(excelFromFile)
        excelTo = excel(excelToFile)
        # 3.加载excel
        ret = excelFrom.open('Sheet1')
        if not ret:
            break
        ret = excelTo.open('Sheet1')
        if not ret:
            break
        logging.debug("load excel %s from and to success" %('sheet1'))
        excelTo.Combine(excelFrom, referColumn, combineColumn)
        # 执行到最后也得退出
        break

    if excelFrom.active:
        excelFrom.close()
    if excelTo.active:
        excelTo.close()
    logging.debug("end autocombine excel".center(30, '='))


if __name__ != '__main__':
    for i in range(10):
        print(i)
